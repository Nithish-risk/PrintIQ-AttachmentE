# filepath: c:\Users\kumarn40\OneDrive - Reed Elsevier Group ICO Reed Elsevier Inc\Desktop\Gen AI\VITALIQ_printiq_data\V 2.1\modules\excel_image_extractor.py
"""Extract images embedded in Excel cells and map them to their source rows.

pandas/openpyxl read only cell *text*; embedded pictures live in the worksheet
drawing layer and are anchored to a cell via ``image.anchor``. This module reads
those anchors (the authoritative image→cell mapping Azure DI cannot provide),
saves each image, and returns a ``{(sheet_row_0based) -> [image_path, ...]}``
map plus the anchor column so the caller can associate an image with the correct
``example`` column of a specific rule row.

The returned row index is **0-based over the raw worksheet rows**, matching the
index space used by ``excel_parser`` (which reads with ``header=None``).
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


def _anchor_cell(image) -> Optional[Tuple[int, int]]:
    """Return (row0, col0) of an image's top-left anchor, or None.

    Handles openpyxl's OneCellAnchor / TwoCellAnchor (both expose ``_from`` with
    0-based ``row`` / ``col``). A string anchor (e.g. "A1") is unusual for images
    but handled defensively.
    """
    anchor = getattr(image, "anchor", None)
    if anchor is None:
        return None
    frm = getattr(anchor, "_from", None)
    if frm is not None:
        row = getattr(frm, "row", None)
        col = getattr(frm, "col", None)
        if row is not None and col is not None:
            return int(row), int(col)
    return None


def extract_cell_images(
    xlsx_path: str | Path,
    sheet: str,
    out_dir: str | Path,
) -> Dict[int, List[dict]]:
    """Return ``{row0: [{"path", "col"}, ...]}`` for images on *sheet*.

    Each entry records the saved image path and the 0-based anchor column so the
    caller can confirm the picture sits in the ``example`` column. Returns an
    empty dict when openpyxl is unavailable or the sheet has no images.
    """
    if load_workbook is None:
        return {}
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        return {}
    if sheet not in wb.sheetnames:
        return {}
    ws = wb[sheet]

    images = getattr(ws, "_images", None) or []
    result: Dict[int, List[dict]] = {}
    for i, image in enumerate(images):
        cell = _anchor_cell(image)
        if cell is None:
            continue
        row0, col0 = cell
        # Image bytes: openpyxl exposes them via ``image.ref`` (a BytesIO) or
        # ``image._data()`` depending on version.
        data = None
        try:
            data = image._data()  # type: ignore[attr-defined]
        except Exception:
            ref = getattr(image, "ref", None)
            if hasattr(ref, "getvalue"):
                data = ref.getvalue()
        if not data:
            continue
        ext = (getattr(image, "format", None) or "png").lower()
        path = out_dir / f"{sheet}_r{row0}_c{col0}_{i}.{ext}"
        try:
            path.write_bytes(data)
        except Exception:
            continue
        result.setdefault(row0, []).append({"path": str(path), "col": col0})
    return result
